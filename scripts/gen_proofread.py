# -*- coding: utf-8 -*-
"""
Build the English source-proofreading workbook for decobesthome from messages/en.json.
Mirrors the viveros_proofread_en format:
  Sheet 1 "Read me first"  - instructions
  Sheet 2 "Strings"        - Key | Section | English source | Current text |
                             Corrected English | Notes / variables | Comments
Pre-fills C/D/E with the live text; column E is where proofreaders edit.
Edited rows (D != E) auto-highlight yellow via conditional formatting.
"""
import json, re, sys, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SRC = "messages/en.json"
OUT = sys.argv[1] if len(sys.argv) > 1 else "decobesthome_proofread_en.xlsx"

en = json.load(open(SRC, encoding="utf-8"))

# ---- flatten preserving insertion order ----------------------------------
rows = []
def walk(o, path):
    if isinstance(o, dict):
        for k, v in o.items():
            walk(v, path + [k])
    elif isinstance(o, list):
        for i, v in enumerate(o):
            walk(v, path + [str(i)])
    else:
        rows.append((".".join(path), o))
walk(en, [])
rows = [(k, v) for k, v in rows if isinstance(v, str)]

# ---- lookups for friendly section names ----------------------------------
prod_names = en["products"]["names"]
post_titles = {pid: (p.get("title") or pid) for pid, p in en["journal"]["posts"].items()}

def is_seo(key):
    return key.startswith("meta.") or re.search(r"\.meta\.", key) is not None

def section_for(key):
    if is_seo(key):
        return "SEO — page titles & meta descriptions"
    if key.startswith("nav."):
        return "Navigation"
    if key.startswith("header."):
        return "Header & mega menu"
    if key.startswith("footer."):
        return "Footer"
    if key.startswith("quote."):
        return "Quote / request form"
    if key.startswith("products.names"):
        return "Products — range names"
    if key.startswith("products.tags"):
        return "Products — tags"
    if key.startswith("products.mega"):
        return "Products — mega-menu descriptors"
    if key.startswith("products.index"):
        return "Products — catalogue page"
    m = re.match(r"products\.detail\.items\.([^.]+)\.", key)
    if m:
        pid = m.group(1)
        return "Product — " + prod_names.get(pid, pid)
    if key.startswith("products.detail"):
        return "Product detail — shared labels"
    if key.startswith("home."):
        return "Home page"
    if key.startswith("about."):
        return "About page"
    m = re.match(r"journal\.posts\.([^.]+)\.", key)
    if m:
        pid = m.group(1)
        return "Journal post — " + post_titles.get(pid, pid)
    if key.startswith("journal.detail"):
        return "Journal — article page labels"
    if key.startswith("journal."):
        return "Journal — index page"
    if key.startswith("inspiration."):
        return "Inspiration page"
    if key.startswith("contact."):
        return "Contact page"
    return key.split(".")[0]

# ---- notes / variables ----------------------------------------------------
# Proper nouns / brand / botanical names worth preserving. Flagged only when present.
PROPER = [
    "BestHome", "Xiamen", "Imp. & Exp.", "Tonkin", "Moso", "Phyllostachys",
    "Phragmites", "Arundo", "Sode-Gaki", "Sode Gaki", "Koetsu", "Kenninji",
    "Yotsume", "Misu", "Teppo", "Kingfisher", "Jiangxi", "Fujian", "Anji",
    "Gui bamboo", "China", "Japanese",
]

def notes_for(key, val):
    parts = []
    ph = re.findall(r"\{[^}]+\}", val)
    if ph:
        parts.append("Variables (keep unchanged): " + ", ".join(dict.fromkeys(ph)))
    if is_seo(key):
        n = len(val)
        if key.endswith(".title") or key == "meta.title" or key == "meta.brand":
            parts.append("SEO title — keep ≤ 60 chars (current: %d)" % n)
        elif key.endswith(".description"):
            parts.append("Meta description — keep ≤ 155 chars (current: %d)" % n)
    found = []
    for name in PROPER:
        if name in val and name not in found:
            found.append(name)
    if found:
        parts.append("⚠ Brand / proper names: " + ", ".join(found))
    return " · ".join(parts)

# ---- order: SEO block first, then natural order --------------------------
seo_rows = [r for r in rows if is_seo(r[0])]
rest_rows = [r for r in rows if not is_seo(r[0])]
ordered = seo_rows + rest_rows

# ---- build workbook -------------------------------------------------------
wb = openpyxl.Workbook()

# ===== Read me first =====
rm = wb.active
rm.title = "Read me first"
rm.column_dimensions["A"].width = 110
GREEN = "FF1E3D32"
readme = [
    ("decobesthome — Source Proofreading · English (the site is English-only)", 14, True),
    ("", 11, False),
    ("What this file is", 12, True),
    ("This is the master copy of every piece of text shown on the decobesthome.com website,", 11, False),
    ("pulled straight from the site's content file (messages/en.json). The site is English-only,", 11, False),
    ("so this single workbook is the complete source of truth for the wording.", 11, False),
    ("", 11, False),
    ("Use this workbook to proofread and refine the copy: fix typos, tighten sentences, sharpen", 11, False),
    ("the tone, and polish the SEO titles & meta descriptions. Hand the file back when done and", 11, False),
    ("the corrections get applied to the website automatically.", 11, False),
    ("", 11, False),
    ("How to use this file", 12, True),
    ("1.  Open the 'Strings' sheet. Each row is one piece of text shown on the website.", 11, False),
    ("2.  Read column C ('English source') — this is the current live text on the site.", 11, False),
    ("    Column D ('Current text') is the same value, kept so the import script can tell", 11, False),
    ("    which rows you changed (it compares D against E).", 11, False),
    ("3.  If the wording is fine, leave column E ('Corrected English') exactly as it is —", 11, False),
    ("    it is already pre-filled with the current text, so no action is needed.", 11, False),
    ("4.  To improve the wording, just edit column E in place. Edited rows turn yellow", 11, False),
    ("    automatically, so you can see at a glance what you changed.", 11, False),
    ("5.  Column F ('Notes / variables') flags things to keep intact: placeholders like", 11, False),
    ("    {email} or {year}, brand/proper names (BestHome, Tonkin, Moso…) and SEO length", 11, False),
    ("    limits. Please respect these when rewording.", 11, False),
    ("6.  Column G ('Comments') is yours — leave any remarks, alternatives or questions.", 11, False),
    ("7.  When finished, save the file and send it back. Do NOT change column A (Key) or", 11, False),
    ("    rename the sheets — rows are matched by Key, not by position, so you may sort/filter freely.", 11, False),
    ("", 11, False),
    ("Things to keep an eye on", 12, True),
    ("•  Brand & company names — BestHome, Xiamen BestHome Imp. & Exp. Co., Ltd.", 11, False),
    ("•  Product / botanical names — Tonkin, Moso, Phyllostachys, Phragmites (reed); and the", 11, False),
    ("   Japanese fence styles (Sode-Gaki, Koetsu, Kenninji, Yotsume…). Keep spellings exact.", 11, False),
    ("•  Placeholders — anything in curly braces, e.g. {email}, {phone}, {year}, {mins}, must", 11, False),
    ("   appear unchanged in the corrected text. Missing/extra placeholders are rejected on import.", 11, False),
    ("•  SEO titles ≤ 60 characters; meta descriptions ≤ 155 characters (otherwise Google truncates).", 11, False),
    ("•  Tone — BestHome is a B2B wholesale supplier of natural bamboo, reed, willow and thatch", 11, False),
    ("   products, addressing trade buyers, landscapers and distributors. Warm but professional;", 11, False),
    ("   confident, not salesy.", 11, False),
    ("", 11, False),
    ("File version: 2026-06-09.   Total rows in the Strings sheet: %d." % len(ordered), 11, False),
]
for i, (txt, sz, bold) in enumerate(readme, start=1):
    c = rm.cell(i, 1, txt)
    c.font = Font(size=sz, bold=bold, color=(GREEN if bold else None))
    c.alignment = Alignment(wrap_text=False, vertical="center")

# ===== Strings =====
ws = wb.create_sheet("Strings")
headers = ["Key", "Section", "English source", "Current text",
           "Corrected English (master copy)", "Notes / variables", "Comments"]
widths = [38, 30, 55, 55, 55, 40, 30]
for c, (h, w) in enumerate(zip(headers, widths), start=1):
    cell = ws.cell(1, c, h)
    cell.fill = PatternFill("solid", fgColor=GREEN)
    cell.font = Font(bold=True, color="FFFFFFFF")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 28

thin = Side(style="thin", color="FFD9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
top_wrap = Alignment(wrap_text=True, vertical="top")

for i, (key, val) in enumerate(ordered, start=2):
    sec = section_for(key)
    note = notes_for(key, val)
    data = [key, sec, val, val, val, note, ""]
    for c, v in enumerate(data, start=1):
        cell = ws.cell(i, c, v)
        cell.alignment = top_wrap
        cell.border = border
    # key + section a touch muted
    ws.cell(i, 1).font = Font(size=9, color="FF666666")
    ws.cell(i, 2).font = Font(size=10, color="FF444444")

ws.freeze_panes = "C2"  # keep Key + Section pinned, header row frozen

# highlight whole row yellow when corrected (E) differs from current (D)
yellow = PatternFill("solid", fgColor="FFFFF2CC")
last = ws.max_row
ws.conditional_formatting.add(
    "A2:G%d" % last,
    FormulaRule(formula=["AND($D2<>\"\",EXACT($D2,$E2)=FALSE)"], fill=yellow),
)

# autofilter
ws.auto_filter.ref = "A1:G%d" % last

wb.save(OUT)
print("Wrote %s  (%d strings, %d SEO rows)" % (OUT, len(ordered), len(seo_rows)))
# quick section tally
from collections import Counter
tally = Counter(section_for(k) for k, _ in ordered)
for s, n in tally.most_common():
    print("  %4d  %s" % (n, s))
