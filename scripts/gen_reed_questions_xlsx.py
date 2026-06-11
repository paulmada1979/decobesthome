# -*- coding: utf-8 -*-
"""Reed-fence supplier questionnaire as an Excel workbook with answer columns."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else "reed-fence-supplier-questions.xlsx"
GREEN = "FF1E3D32"; LEAF = "FF84BE45"; BONE = "FFF6F3EC"
ANSWER = "FFFFFDF5"      # very light cream = "type here"
MUTE = "FF666666"

wb = openpyxl.Workbook()

thin = Side(style="thin", color="FFD9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top")
wrap_ctr = Alignment(wrap_text=True, vertical="center")

def header_cell(c, text):
    c.value = text; c.fill = PatternFill("solid", fgColor=GREEN)
    c.font = Font(bold=True, color="FFFFFFFF", size=10.5)
    c.alignment = wrap_ctr; c.border = border

VERSIONS = [
    "Extra-thick peeled (8–10 mm)",
    "Thick peeled (5–8 mm)",
    "High-density fine (3–6 mm)",
    "Unpeeled thick reed",
    "Half-round / split reed",
    "Japanese-style with bamboo",
]

# ============ Sheet 1: Read me ============
rm = wb.active; rm.title = "Read me first"; rm.column_dimensions["A"].width = 112
lines = [
    ("Reed fence series — supplier questionnaire", 14, True),
    ("", 11, False),
    ("We are building a separate web page for each of these 6 reed-fence versions. To describe each", 11, False),
    ("one factually (and answer the questions buyers search for), we need your input.", 11, False),
    ("", 11, False),
    ("How to fill this in", 12, True),
    ("• Sheet 'Specs per version' — each ROW is one detail; fill an answer in the column for each of", 11, False),
    ("  the 6 versions. Numbers/ranges where possible; rough estimates are fine (just mark them ~).", 11, False),
    ("• Sheet 'Questions & answers' — type your reply in the green-headed 'Answer' column on the right.", 11, False),
    ("• Anything you don't know, leave blank — partial is better than nothing.", 11, False),
    ("", 11, False),
    ("Most important for us", 12, True),
    ("• Opacity / privacy — can you see through it? roughly how 'closed' is it (gaps between stems)?", 11, False),
    ("• Lifespan outdoors and how the colour changes over the years.", 11, False),
    ("• Why a customer would choose each version over the others (the single biggest reason).", 11, False),
    ("", 11, False),
    ("The 6 versions: 1) Extra-thick peeled 8–10 mm  2) Thick peeled 5–8 mm  3) High-density fine 3–6 mm", 11, False),
    ("4) Unpeeled thick reed  5) Half-round / split reed  6) Japanese-style bound with bamboo.", 11, False),
]
for i, (t, sz, b) in enumerate(lines, 1):
    c = rm.cell(i, 1, t); c.font = Font(size=sz, bold=b, color=(GREEN if b else None))
    c.alignment = Alignment(vertical="center")

# ============ Sheet 2: Specs per version (grid) ============
sp = wb.create_sheet("Specs per version")
# header row
header_cell(sp.cell(1, 1), "Detail / spec")
for j, v in enumerate(VERSIONS, start=2):
    header_cell(sp.cell(1, j), v)
sp.column_dimensions["A"].width = 34
for j in range(2, 2 + len(VERSIONS)):
    sp.column_dimensions[get_column_letter(j)].width = 22
sp.row_dimensions[1].height = 40
sp.freeze_panes = "B2"

specs = [
    "In one line: what is this version & what makes it unique?",
    "Item / model code",
    "Reed species & origin (region, harvest season)",
    "Reed diameter (mm) — confirm",
    "Peeled / unpeeled / split",
    "Density — reeds per metre OR kg per m²",
    "Opacity / privacy — see through it? approx % closed? gaps?",
    "Wind — blocks wind, or lets it pass?",
    "Wires — material (galvanised?), no. of rows, gauge",
    "Binding — single/double wire; machine or hand-bound",
    "One-sided or two-sided 'good' face",
    "Heights available (cm)",
    "Roll length (m)",
    "Weight — per roll / per m²",
    "Lifespan outdoors (years, normal EU garden)",
    "How the colour ages (greys/fades? how fast? to what?)",
    "Treatment (oil / coating / anti-mould / none)",
    "Best use cases (garden, balcony, pool, terrace, commercial)",
    "Price tier — budget / mid / premium (+ rough index)",
    "MOQ (minimum order)",
    "Packing — pcs per pallet, pcs per 20ft & 40ft container",
    "Custom options — heights/lengths, private label?",
    "Certificates (FSC, fumigation, phytosanitary…)",
    "WHY choose this one — the single biggest reason",
    "Main trade-off vs the neighbouring versions",
    "Who is it NOT right for?",
]
for r, label in enumerate(specs, start=2):
    c = sp.cell(r, 1, label)
    c.alignment = wrap_top; c.border = border
    c.fill = PatternFill("solid", fgColor=BONE); c.font = Font(size=9.5)
    for j in range(2, 2 + len(VERSIONS)):
        a = sp.cell(r, j); a.fill = PatternFill("solid", fgColor=ANSWER)
        a.alignment = wrap_top; a.border = border
    sp.row_dimensions[r].height = 30

# ============ Sheet 3: Questions & answers (two-column) ============
qa = wb.create_sheet("Questions & answers")
qa.column_dimensions["A"].width = 14
qa.column_dimensions["B"].width = 78
qa.column_dimensions["C"].width = 60
header_cell(qa.cell(1, 1), "Part")
header_cell(qa.cell(1, 2), "Question")
header_cell(qa.cell(1, 3), "Answer")
qa.row_dimensions[1].height = 26
qa.freeze_panes = "A2"

def section(ws, r, text):
    c = ws.cell(r, 1, text)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c.fill = PatternFill("solid", fgColor=LEAF)
    c.font = Font(bold=True, color=GREEN, size=10.5)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 22

def qrow(ws, r, part, q):
    p = ws.cell(r, 1, part); p.alignment = wrap_top; p.border = border
    p.font = Font(size=9, color=MUTE)
    qc = ws.cell(r, 2, q); qc.alignment = wrap_top; qc.border = border; qc.font = Font(size=10)
    qc.fill = PatternFill("solid", fgColor=BONE)
    a = ws.cell(r, 3, ""); a.alignment = wrap_top; a.border = border
    a.fill = PatternFill("solid", fgColor=ANSWER)
    ws.row_dimensions[r].height = 34

r = 2
section(qa, r, "Per-version clarifiers (the small differences buyers ask about)"); r += 1
clarifiers = [
    "Extra-thick peeled (8–10 mm): is this the most private / 'most closed' option? the heaviest? the longest-lasting?",
    "Thick peeled (5–8 mm): is this the everyday best-seller / default choice — and why?",
    "High-density fine (3–6 mm): is choosing this about a smoother, more refined look rather than privacy? Is it more or less closed than the thick peeled?",
    "Unpeeled thick reed: what does keeping the skin on change in practice — look, lifespan, mould risk, smell, price? Cheaper or dearer than peeled?",
    "Half-round / split reed: why split the reed — flatter face, less depth, lighter shipping, lower price? Is it as private as round reed?",
    "Japanese-style with bamboo: what bamboo is used and what's the binding pattern? Mainly decorative, or also full privacy? Premium-priced?",
]
for q in clarifiers:
    qrow(qa, r, "B", q); r += 1

section(qa, r, "Series-level questions (answer once for the whole range)"); r += 1
series = [
    "Which version is your best-seller? Which is the cheapest, and which the most premium?",
    "Which is the most 'closed' — no gaps, maximum privacy — and which lets the most light/air through?",
    "Peeled vs unpeeled in real terms: difference in lifespan, mould resistance, smell, and how the colour changes over the years?",
    "Which versions are safe near a swimming pool / on the coast (salt, chlorine, constant damp)?",
    "How is each one fixed to a fence or railing (wire ties, battens)? Any difference between versions?",
    "Is any version heavier/denser and therefore needs a stronger supporting frame?",
]
for q in series:
    qrow(qa, r, "C", q); r += 1

wb.save(OUT)
print("Wrote", OUT, "—", len(specs), "spec rows ×", len(VERSIONS), "versions +",
      len(clarifiers) + len(series), "questions")
